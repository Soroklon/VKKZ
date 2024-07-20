import sqlite3
import bottle
import os
import glob
import traceback
import time
import datetime
from bottle import route, run, template, request, response, abort, error, debug, static_file, redirect, HTTPResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
import openpyxl.styles.numbers

HOST = 'localhost'
PORT = 8080

@route('/<filename:path>')
def static(filename):
    return static_file(filename, root='./')

@route('/home')
def startPage():  
    return template('index')

def Translite(ruStr):
    ruLetters = ['А','Б','В','Г','Д','Е','Ё','Ж','З','И','Й','К','Л','М','Н','О','П','Р','С','Т','У','Ф','Х','Ц','Ч','Ш','Щ','Ъ','Ы','Ь','Э','Ю','Я']
    ruStrUp = ruStr.upper()
    latStr = ''
    i = 0
    convert = False
    while(i < len(ruStrUp)):
        s = ruStrUp[i]
        if s in ruLetters:
            if not convert:
                convert = True
                latStr = latStr + '{'
            match s:
                case 'А':
                    latStr = latStr + 'A'
                case 'Б':
                    latStr = latStr + 'B'
                case 'В':
                    latStr = latStr + 'V'
                case 'Г':
                    latStr = latStr + 'G'
                case 'Д':
                    latStr = latStr + 'D'
                case 'Е':
                    latStr = latStr + 'E'
                case 'Ё':
                    latStr = latStr + 'JE'
                case 'Ж':
                    latStr = latStr + 'ZH'
                case 'З':
                    latStr = latStr + 'Z'
                case 'И':
                    latStr = latStr + 'I'
                case 'Й':
                    latStr = latStr + 'Y'
                case 'К':
                    latStr = latStr + 'K'
                case 'Л':
                    latStr = latStr + 'L'
                case 'М':
                    latStr = latStr + 'M'
                case 'Н':
                    latStr = latStr + 'N'
                case 'О':
                    latStr = latStr + 'O'
                case 'П':
                    latStr = latStr + 'P'
                case 'Р':
                    latStr = latStr + 'R'
                case 'С':
                    latStr = latStr + 'S'
                case 'Т':
                    latStr = latStr + 'T'
                case 'У':
                    latStr = latStr + 'U'
                case 'Ф':
                    latStr = latStr + 'F'
                case 'Х':
                    latStr = latStr + 'KH'
                case 'Ц':
                    latStr = latStr + 'C'
                case 'Ч':
                    latStr = latStr + 'CH'
                case 'Ш':
                    latStr = latStr + 'SH'
                case 'Щ':
                    latStr = latStr + 'JSH'
                case 'Ъ':
                    latStr = latStr + 'HH'
                case 'Ы':
                    latStr = latStr + 'IH'
                case 'Ь':
                    latStr = latStr + 'JH'
                case 'Э':
                    latStr = latStr + 'EH'
                case 'Ю':
                    latStr = latStr + 'JU'
                case 'Я':
                    latStr = latStr + 'JA'
        else:
            if convert:
                convert = False
                latStr = latStr + '}'
            latStr = latStr + s
        i += 1
    return latStr

def UnTranslite(latStr):
    latStrUp = latStr.upper()
    ruStr = ''
    i = 0
    convert = False
    while(i < len(latStrUp)):
        s = latStrUp[i]
        if s == '{':
            convert = True
            i += 1
            continue
        if s == '}':
            convert = False
            i += 1
            continue
        if convert:
            if (s == 'J'):
                i += 1
                s = latStrUp[i]
                match s:
                    case 'E':
                        ruStr = ruStr + 'Ё'
                    case 'S':
                        ruStr = ruStr + 'Щ'
                        i += 1
                    case 'H':
                        ruStr = ruStr + 'Ь'
                    case 'U':
                        ruStr = ruStr + 'Ю'
                    case 'A':
                        ruStr = ruStr + 'Я'
            
            elif (i + 1 < len(latStrUp)) and (latStrUp[i+1] == 'H') and not((i + 2 < len(latStrUp)) and (latStrUp[i+2] == 'H')):
                match s:
                    case 'Z':
                        ruStr = ruStr + 'Ж'
                    case 'K':
                        ruStr = ruStr + 'Х'
                    case 'C':
                        ruStr = ruStr + 'Ч'
                    case 'S':
                        ruStr = ruStr + 'Ш'
                    case 'E':
                        ruStr = ruStr + 'Э'
                    case 'H':
                        ruStr = ruStr + 'Ъ'
                    case 'I':
                        ruStr = ruStr + 'Ы'
                i += 1
            else:
                match s:
                    case 'A':
                        ruStr = ruStr + 'А'
                    case 'B':
                        ruStr = ruStr + 'Б'
                    case 'V':
                        ruStr = ruStr + 'В'
                    case 'G':
                        ruStr = ruStr + 'Г'
                    case 'D':
                        ruStr = ruStr + 'Д'
                    case 'E':
                        ruStr = ruStr + 'Е'
                    case 'Z':
                        ruStr = ruStr + 'З'
                    case 'I':
                        ruStr = ruStr + 'И'
                    case 'Y':
                        ruStr = ruStr + 'Й'
                    case 'K':
                        ruStr = ruStr + 'К'
                    case 'L':
                        ruStr = ruStr + 'Л'
                    case 'M':
                        ruStr = ruStr + 'М'
                    case 'N':
                        ruStr = ruStr + 'Н'
                    case 'O':
                        ruStr = ruStr + 'О'
                    case 'P':
                        ruStr = ruStr + 'П'
                    case 'R':
                        ruStr = ruStr + 'Р'
                    case 'S':
                        ruStr = ruStr + 'С'
                    case 'T':
                        ruStr = ruStr + 'Т'
                    case 'U':
                        ruStr = ruStr + 'У'
                    case 'F':
                        ruStr = ruStr + 'Ф'
                    case 'C':
                        ruStr = ruStr + 'Ц'
        else:
             ruStr = ruStr + s   
        i += 1
    return ruStr

@route('/sidebar')
def startPage():
    cResult = []
    rResult = []
    cResultTrans = []
    rResultTrans = []
    if os.path.exists('ReportsDB.db'):
        conn = sqlite3.connect('ReportsDB.db', uri=True)
        try:
            c = conn.cursor()
            c.execute("SELECT ComponentID, ComponentName FROM Components UNION SELECT ComponentID, ComponentName FROM ComponentsIncoming ORDER BY ComponentID")
            cResult = c.fetchall()
            c.close()
            c = conn.cursor()
            c.execute("SELECT DISTINCT RecipeID, RecipeName FROM Recipes ORDER BY RecipeID")
            rResult = c.fetchall()
            c.close()
            for i in cResult:
                cResultTrans.append([i[0],UnTranslite(i[1])])
            for i in rResult:
                rResultTrans.append([i[0],UnTranslite(i[1])])
        except sqlite3.Error:
            cResult = []
            rResult = []
            cResultTrans = []
            rResultTrans = []
        finally:
            c.close()
    return template('sidebar', cList=cResultTrans, rList=rResultTrans)

@route('/startPage')
def startPage():
    status = False
    if os.path.exists('ReportsDB.db'):
        status = True
    return template('startPage', ans = status)

@route('/recipeReq', method='GET')
def recipeReq():
    stamp = datetime.datetime.now()
    tStart = str(request.GET.get('FROM'))
    tEnd = str(request.GET.get('TO'))
    rId = str(request.GET.get('REC'))
    tStartStr = "<не задано>"
    tEndStr = "<не задано>"
    files = glob.glob('ReportsFolder/*')
    for f in files:
        os.remove(f)
    wb = Workbook()
    ws= wb.active
    
    anyParam = False
    cols_dict = {}
    if os.path.exists('ReportsDB.db'):
        conn = sqlite3.connect('ReportsDB.db', uri=True)
    try:
    #Запрос в общую таблицу по заданным параметрам
        c = conn.cursor()
        reqString = "SELECT * FROM Recipes JOIN Components ON Components.TimeStmp = Recipes.TimeStmp WHERE"
        if tStart != "0":
            reqString = reqString + " (Recipes.TimeStmp >= " + tStart + ")"
            tStartStr = datetime.datetime.fromtimestamp(float(tStart)/1000.0).strftime('%d.%m.%Y %H:%M:%S')
            anyParam = True
        if tEnd != "0":
            if anyParam:
                reqString = reqString + " AND"
            reqString = reqString + " (Recipes.TimeStmp <= " + tEnd + ")"
            tEndStr = datetime.datetime.fromtimestamp(float(tEnd)/1000.0).strftime('%d.%m.%Y %H:%M:%S')
            anyParam = True
        if rId != "0":
            if anyParam:
                reqString = reqString + " AND"
            reqString = reqString+ " (Recipes.RecipeID IN (" + rId + "))"
        reqString = reqString + " ORDER BY Recipes.TimeStmp"
        c.execute(reqString)
        result = c.fetchall()
        c.close()
    #Разбор на отдельные массивы с рецептами и компонентами
        rAns = []
        cTemp = []
        cAns = []
        
        tStmp = 0
        for r in result:
            if tStmp != r[4]:
                tStmp = r[4]
                rAns.append([r[0],UnTranslite(r[1]),round(abs(r[2]),2),round(abs(r[3]),2),datetime.datetime.fromtimestamp(float(r[4])/1000.0).strftime('%d.%m.%Y %H:%M:%S')])
                if cTemp:
                    cAns.append(cTemp)
                    cTemp = []
            cTemp.append([r[5],UnTranslite(r[6]),round(abs(r[7]),2),round(abs(r[8]),2)])
        if cTemp:
            cAns.append(cTemp)
            cTemp = []
    #Подсчёт суммарных план/факт
        rTotal = []
        cTotal = []
        cTemp = []

        if rAns:
            wInd = 0
            rInd = 0
            first = True
            exist = False

            for rAnsRow in rAns:
                if first:
                    rTotal.append([rAnsRow[0], rAnsRow[1], 0.0, 0.0])
                    for cAnsRow in cAns[rInd]:
                        cTemp.append([cAnsRow[0], cAnsRow[1], 0.0, 0.0])
                    cTotal.append(cTemp)
                    cTemp = []
                    first = False
                if (rAnsRow[0] != rTotal[wInd][0]):
                    exist = False
                    i = 0
                    for fRows in rTotal:
                        if (rAns[0] == fRows[0]):
                            exist = True
                            wInd = i
                            break
                        i = i + 1
                    if not exist:
                        wInd = i
                        rTotal.append([rAnsRow[0], rAnsRow[1], 0.0, 0.0])
                        for cAnsRow in cAns[rInd]:
                            cTemp.append([cAnsRow[0], cAnsRow[1], 0.0, 0.0])
                        cTotal.append(cTemp)
                        cTemp = []
                rTotal[wInd][2] = round(rTotal[wInd][2] + rAnsRow[2], 2)
                rTotal[wInd][3] = round(rTotal[wInd][3] + rAnsRow[3], 2)
                i = 0
                for cAnsRow in cAns[rInd]:
    #added
                    j = 0
                    exist = False
                    for cFindRow in cTotal[wInd]:
                        if (cTotal[wInd][j][0] == cAnsRow[0]):
                            exist = True
                            break
                        j = j + 1
                    if not exist:
                        cTotal[wInd].append([cAnsRow[0], cAnsRow[1], 0.0, 0.0])
    #end
                    cTotal[wInd][j][2] = round(cTotal[wInd][j][2] + cAnsRow[2], 2)
                    cTotal[wInd][j][3] = round(cTotal[wInd][j][3] + cAnsRow[3], 2)
                    i = i + 1
                rInd = rInd + 1
    #------------------Запись в Excel---------------------------
    #------------------Шапка------------------------------------
        ws.merge_cells('A1:D1')
        ws.merge_cells('B2:C2')
        ws.merge_cells('B3:C3')
        ws['A1'].value = "Отчёт по рецептам за период"
        ws['A2'].value = "с:"
        ws['A2'].alignment = Alignment(horizontal='right')
        ws['B2'].value = tStartStr
        ws['A3'].value = "по:"
        ws['A3'].alignment = Alignment(horizontal='right')
        ws['B3'].value = tEndStr
        ws.append([])
        ws['A5'].value = "Суммарно"
        ws.append(['Код рецепта', 'Наименование рецепта', 'Код компонента', 'Наименование компонента', 'План', 'Факт'])
        for col in range(1,7):
            ws.cell(row = 6, column = col).font = Font(bold = True)
        iTotal = 7
        ws.append([])
        ws['A8'].value = "История"
        ws.append(['Код рецепта', 'Наименование рецепта', 'Код компонента', 'Наименование компонента', 'План', 'Факт', 'Дата/время'])
        for col in range(1,8):
            ws.cell(row = 9, column = col).font = Font(bold = True)
        iHistory = 10
    #--------------Заполнение таблицы---------------------------
        k = 0
        for r in rTotal:
            ws.insert_rows(iTotal)
            ws.cell(row = iTotal, column = 1).value = r[0]
            ws.cell(row = iTotal, column = 2).value = r[1]
            ws.cell(row = iTotal, column = 5).value = r[2]
            ws.cell(row = iTotal, column = 5).number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[2]
            ws.cell(row = iTotal, column = 6).value = r[3]
            ws.cell(row = iTotal, column = 6).number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[2]
            iTotal = iTotal + 1
            iHistory = iHistory + 1
            for ci in cTotal[k]:
                ws.insert_rows(iTotal)
                jTotal = 3
                for cij in ci:
                    ws.cell(row = iTotal, column = jTotal).value = cij
                    if (isinstance(cij, float)):
                        ws.cell(row = iTotal, column = jTotal).number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[2]
                    jTotal = jTotal + 1
                iTotal = iTotal + 1
                iHistory = iHistory + 1
            k = k + 1
        k = 0
        for r in rAns:
            ws.cell(row = iHistory, column = 1).value = r[0]
            ws.cell(row = iHistory, column = 2).value = r[1]
            ws.cell(row = iHistory, column = 5).value = r[2]
            ws.cell(row = iHistory, column = 5).number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[2]
            ws.cell(row = iHistory, column = 6).value = r[3]
            ws.cell(row = iHistory, column = 6).number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[2]
            ws.cell(row = iHistory, column = 7).value = r[4]
            iHistory = iHistory + 1
            for ci in cAns[k]:
                jHistory = 3
                for cij in ci:
                    ws.cell(row = iHistory, column = jHistory).value = cij
                    if (isinstance(cij, float)):
                        ws.cell(row = iHistory, column = jHistory).number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[2]
                    jHistory = jHistory + 1
                iHistory = iHistory + 1
            k = k + 1
    #--------------Авторазмер ячеек------------------------------
        for row in ws.rows:
            for cell in row:
                if cell.coordinate in ws.merged_cells:
                    continue
                letter = cell.column_letter
                if cell.value:
                    len_cell = len(str(cell.value))
                    len_cell_dict = 0
                    if letter in cols_dict:
                        len_cell_dict = cols_dict[letter]
                    if len_cell > len_cell_dict:
                        cols_dict[letter] = len_cell
                        new_width_col = len_cell * 1.4
                        ws.column_dimensions[cell.column_letter].width = new_width_col
    except sqlite3.Error:
        code = traceback.format_exc()
        return template('errpage', er = code)
    finally:
            c.close()
            ReportName = 'ReportsFolder/RecipesReport_' + stamp.strftime('%Y%m%d_%H%M%S') + '.xlsx'
            wb.save(filename=ReportName)
    return template('recipeReq', rl=rAns, cl=cAns, rt=rTotal, ct=cTotal, tFrom = tStartStr, tTo = tEndStr, rName = ReportName)

@route('/componentReq', method='GET')
def componentReq():
    stamp = datetime.datetime.now()
    tStart = str(request.GET.get('FROM'))
    tEnd = str(request.GET.get('TO'))
    cId = str(request.GET.get('COMP'))
    tStartStr = "<не задано>"
    tEndStr = "<не задано>"
    files = glob.glob('ReportsFolder/*')
    for f in files:
        os.remove(f)
    wb = Workbook()
    ws = wb.active
    
    anyParam = False
    cols_dict = {}
    if os.path.exists('ReportsDB.db'):
        conn = sqlite3.connect('ReportsDB.db', uri=True)
    try:
    #Запрос в общую таблицу по заданным параметрам
        c = conn.cursor()
        reqString = "SELECT * FROM (SELECT ComponentID, ComponentName, FactWeight, TimeStmp	FROM Components UNION SELECT ComponentID, ComponentName, FactWeight, TimeStmp FROM ComponentsIncoming) WHERE"
        if tStart != "0":
            reqString = reqString + " (TimeStmp >= " + tStart + ")"
            tStartStr = datetime.datetime.fromtimestamp(float(tStart)/1000.0).strftime('%d.%m.%Y %H:%M:%S')
            anyParam = True
        if tEnd != "0":
            if anyParam:
                reqString = reqString + " AND"
            reqString = reqString + " (TimeStmp <= " + tEnd + ")"
            tEndStr = datetime.datetime.fromtimestamp(float(tEnd)/1000.0).strftime('%d.%m.%Y %H:%M:%S')
            anyParam = True
        if cId != "0":
            if anyParam:
                reqString = reqString + " AND"
            reqString = reqString + " (ComponentID IN (" + cId + "))"
        reqString = reqString + " ORDER BY TimeStmp"
        c.execute(reqString)
        result = c.fetchall()
        c.close()
    #Разбор на массивы по ID компонентов + массив с суммарными приход/расход
        cAns = []
        cTemp = []
        cTotal = []

        cId = 0
        plus = 2
        minus = 3
        if result:
            wInd = 0
            rInd = 0
            first = True
            exist = False
            
            for r in result:
                if first:
                    cTotal.append([r[0], UnTranslite(r[1]), 0.0, 0.0])
                    cAns.append([])
                    first = False
                if (cTotal[wInd][0] != r[0]):
                    exist = False
                    i = 0
                    for fRow in cTotal:
                        if (fRow[0] == r[0]):
                            exist = True
                            wInd = i
                            break
                        i = i + 1
                    if not exist:
                        wInd = i
                        cTotal.append([r[0], UnTranslite(r[1]), 0.0, 0.0])
                        cAns.append([])
                if (r[2] > 0):
                    cAns[wInd].append([abs(r[2]), 0.0, datetime.datetime.fromtimestamp(float(r[3])/1000.0).strftime('%d.%m.%Y %H:%M:%S')])
                    cTotal[wInd][plus] = round((cTotal[wInd][plus] + abs(r[2])), 2)
                else:
                    cAns[wInd].append([0.0, abs(r[2]), datetime.datetime.fromtimestamp(float(r[3])/1000.0).strftime('%d.%m.%Y %H:%M:%S')])
                    cTotal[wInd][minus] = round((cTotal[wInd][minus] + abs(r[2])), 2)
    #------------------Запись в Excel---------------------------
    #------------------Шапка------------------------------------
            ws.merge_cells('A1:D1')
            ws.merge_cells('B2:C2')
            ws.merge_cells('B3:C3')
            ws['A1'].value = "Отчёт по компонентам за период"
            ws['A2'].value = "с:"
            ws['A2'].alignment = Alignment(horizontal='right')
            ws['B2'].value = tStartStr
            ws['A3'].value = "по:"
            ws['A3'].alignment = Alignment(horizontal='right')
            ws['B3'].value = tEndStr
            ws.append([])
            #ws.merge_cells('A5:B5')
            ws['A5'].value = "Суммарно"
            ws.append(['Код', 'Наименование', 'Приход', 'Расход'])
            for col in range(1,5):
                ws.cell(row = 6, column = col).font = Font(bold = True)
            iTotal = 7
            ws.append([])
            #ws.merge_cells('A9:B9')
            ws['A8'].value = "История"
            ws.append(['Код', 'Наименование', 'Приход', 'Расход', 'Дата/время'])
            for col in range(1,6):
                ws.cell(row = 9, column = col).font = Font(bold = True)
            iHistory = 10
   #------------------Заполнение таблицы------------------------         
            k = 0
            for r in cTotal:
                ws.insert_rows(iTotal)
                jTotal = 1
                for ri in r:
                    ws.cell(row = iTotal, column = jTotal).value = ri
                    if (isinstance(ri, float)):
                        ws.cell(row = iTotal, column = jTotal).number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[2]
                    jTotal = jTotal + 1
                iTotal = iTotal + 1
                iHistory = iHistory + 1
                ws.append([r[0], r[1]])
                for ri in cAns[k]:
                    jHistory = 3
                    for rij in ri:
                        ws.cell(row = iHistory, column = jHistory).value = rij
                        if (isinstance(rij, float)):
                            if (rij == 0):
                                ws.cell(row = iHistory, column = jHistory).value = '-'
                                ws.cell(row = iHistory, column = jHistory).alignment = Alignment(horizontal='center')
                            else:
                                ws.cell(row = iHistory, column = jHistory).number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[2]
                        jHistory = jHistory + 1
                    iHistory = iHistory + 1
    #--------------Авторазмер ячеек------------------------------
            for row in ws.rows:
                for cell in row:
                    if cell.coordinate in ws.merged_cells:
                        continue
                    letter = cell.column_letter
                    if cell.value:
                        len_cell = len(str(cell.value))
                        len_cell_dict = 0
                        if letter in cols_dict:
                            len_cell_dict = cols_dict[letter]
                        if len_cell > len_cell_dict:
                            cols_dict[letter] = len_cell
                            new_width_col = len_cell * 1.4
                            ws.column_dimensions[cell.column_letter].width = new_width_col
            
    except sqlite3.Error:
        code = traceback.format_exc()
        return template('errpage', er = code)
    finally:
            c.close()
            ReportName = 'ReportsFolder/ComponentsReport_' + stamp.strftime('%Y%m%d_%H%M%S') + '.xlsx'
            wb.save(filename=ReportName)
    return template('componentReq', cl = cAns, ct = cTotal, tFrom = tStartStr, tTo = tEndStr, rName = ReportName)

@route('/componentSumReq')
def componentSumReq():
    files = glob.glob('ReportsFolder/*')
    for f in files:
        os.remove(f)
    wb = Workbook()
    ws = wb.active
    
    anyParam = False
    cols_dict = {}
    if os.path.exists('ReportsDB.db'):
        conn = sqlite3.connect('ReportsDB.db', uri=True)
    try:
    #Запрос в таблицу
        c = conn.cursor()
        reqString = "SELECT * FROM ComponentsSummary ORDER BY ComponentType"
        c.execute(reqString)
        result = c.fetchall()
        c.close()
    #Разбор на массивы по типу компонентов
        cTypes = []
        cSum = []

        if result:
            wInd = 0
            rInd = 0
            first = True
            exist = False
            stamp = result[0][4]
            tStr = datetime.datetime.fromtimestamp(float(stamp)/1000.0).strftime('%d.%m.%Y %H:%M:%S')
            
            for r in result:
                if first:
                    cTypes.append(r[0])
                    cSum.append([[r[1], UnTranslite(r[2]), round(r[3], 2)]])
                    first = False
                else:
                    if (cTypes[wInd] != r[0]):
                        exist = False
                        i = 0
                        for fRow in cTypes:
                            if (fRow[0] == r[0]):
                                exist = True
                                wInd = i
                                break
                            i = i + 1
                        if not exist:
                            wInd = i
                            cTypes.append(r[0])
                            cSum.append([[r[1], UnTranslite(r[2]), round(r[3], 2)]])
                    else:
                        cSum[wInd].append([r[1], UnTranslite(r[2]), round(r[3], 2)])
            i = 0
            for r in cTypes:
                cTypes[i] = UnTranslite(cTypes[i])
                i += 1
    #------------------Запись в Excel---------------------------
    #------------------Шапка------------------------------------
            ws.merge_cells('A1:C1')
            ws.merge_cells('B2:C2')
            ws['A1'].value = "Отчёт по остаткам компонентов"
            ws['A2'].value = "на:"
            ws['A2'].alignment = Alignment(horizontal='right')
            ws['B2'].value = tStr
            ws.append([])
            iRow = 4
   #------------------Заполнение таблицы------------------------         
            k = 0
            for r in cTypes:
                ws.append(['тип:',r])
                for col in range(1,3):
                    ws.cell(row = iRow, column = col).font = Font(bold = True)
                ws.cell(row = iRow, column = 1).alignment = Alignment(horizontal='right')
                iRow += 1
                ws.append(['Код', 'Наименование', 'Остаток'])
                for col in range(1,4):
                    ws.cell(row = iRow, column = col).font = Font(bold = True)
                iRow += 1
                for ri in cSum[k]:
                    ws.append(ri)
                    ws.cell(row = iRow, column = 3).number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[2]
                    iRow += 1
                k += 1
   #--------------Авторазмер ячеек------------------------------
            for row in ws.rows:
                for cell in row:
                    if cell.coordinate in ws.merged_cells:
                        continue
                    letter = cell.column_letter
                    if cell.value:
                        len_cell = len(str(cell.value))
                        len_cell_dict = 0
                        if letter in cols_dict:
                            len_cell_dict = cols_dict[letter]
                        if len_cell > len_cell_dict:
                            cols_dict[letter] = len_cell
                            new_width_col = len_cell * 1.4
                            ws.column_dimensions[cell.column_letter].width = new_width_col
            
    except sqlite3.Error:
        code = traceback.format_exc()
        return template('errpage', er = code)
    finally:
            c.close()
            ReportName = 'ReportsFolder/ComponentsSummaryReport_' + datetime.datetime.fromtimestamp(float(stamp)/1000.0).strftime('%Y%m%d_%H%M%S') + '.xlsx'
            wb.save(filename=ReportName)
    return template('summaryReq', cT = cTypes, cS = cSum, tFrom = tStr, rName = ReportName)

@route('/login')
def loging():
    return template('login')

@route('/login', method='POST')
def do_loging():
    status = False
    login = request.forms.get('log')
    password = request.forms.get('pass')
    if (login == "Admin") and (password == "1234@Qwe"):
        status = True
    if status:
        conn = sqlite3.connect('ReportsDB.db')
        reqStr = "CREATE TABLE Recipes (RecipeId TEXT, RecipeName TEXT, TargetWeight REAL, FactWeight REAL, TimeStmp INT);"
        c = conn.cursor()
        c.execute(reqStr)
        result = c.fetchall()
        c.close()
        reqStr = "CREATE TABLE Components (ComponentId TEXT, ComponentName TEXT, TargetWeight REAL, FactWeight REAL, TimeStmp INT);"
        c = conn.cursor()
        c.execute(reqStr)
        result = c.fetchall()
        c.close()
        reqStr = "CREATE TABLE ComponentsIncoming (ComponentId TEXT, ComponentName TEXT, FactWeight REAL, TimeStmp INT);"
        c = conn.cursor()
        c.execute(reqStr)
        result = c.fetchall()
        c.close()
        reqStr = "CREATE TABLE ComponentsSummary (ComponentType TEXT, ComponentId TEXT, ComponentName TEXT, SummaryWeight REAL, TimeStmp INT);"
        c = conn.cursor()
        c.execute(reqStr)
        result = c.fetchall()
        c.close()
    return template('do_loging', ok=status)

@bottle.error(404)
def not_found(error):
    code = traceback.format_exc()
    return template('errpage', er = code)

@bottle.error(500)
def not_found(error):
    code = traceback.format_exc()
    return template('errpage', er = code)

run(host=HOST, port=PORT, debug=True)

    
